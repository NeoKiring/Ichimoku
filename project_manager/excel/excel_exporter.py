"""
Excelエクスポーター
プロジェクトデータをExcelファイルに書き込む機能
"""
import os
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import Project, Phase, Process, Task, ProjectStatus, TaskStatus
from .excel_utils import set_cell_value


class ExcelExporter:
    """
    プロジェクトデータをExcelファイルに書き込むクラス
    """
    
    def __init__(self):
        """エクスポーターの初期化"""
        self.template_path = os.path.join(os.path.dirname(__file__), "templates", "project_template.xlsx")
        self._has_template = os.path.exists(self.template_path)
    
    def export_to_file(self, project: Project, file_path: str) -> bool:
        """
        プロジェクトをExcelファイルにエクスポート
        
        Args:
            project: エクスポートするプロジェクト
            file_path: 出力先のファイルパス
            
        Returns:
            エクスポートが成功したかどうか
        """
        try:
            # テンプレートが存在する場合は使用、なければ新規作成
            if self._has_template:
                workbook = openpyxl.load_workbook(self.template_path)
            else:
                workbook = openpyxl.Workbook()
                self._create_default_sheets(workbook)
            
            # スケジュールシートの設定
            self._setup_schedule_sheet(workbook, project)
            
            # 入力シートの設定
            self._setup_input_sheet(workbook, project)
            
            # タスクシートの設定
            self._create_task_sheets(workbook, project)
            
            # ファイルを保存
            workbook.save(file_path)
            return True
            
        except Exception as e:
            print(f"Excelのエクスポート中にエラーが発生しました: {str(e)}")
            return False
    
    def _create_default_sheets(self, workbook):
        """
        デフォルトのシート構造を作成
        
        Args:
            workbook: Excelワークブック
        """
        # デフォルトシートの名前を変更
        default_sheet = workbook.active
        default_sheet.title = "スケジュール"
        
        # 入力シートを追加
        workbook.create_sheet("入力")
    
    def _setup_schedule_sheet(self, workbook, project: Project):
        """
        スケジュールシートを設定
        
        Args:
            workbook: Excelワークブック
            project: プロジェクト
        """
        schedule_sheet = workbook["スケジュール"]
        
        # ヘッダー設定
        set_cell_value(schedule_sheet, "B2", project.name)
        
        # 列ヘッダー
        headers = ["ID", "名前", "担当者", "進捗", "開始日", "終了日", "予想工数", "実工数"]
        for i, header in enumerate(headers):
            col = chr(66 + i)  # B, C, D, ...
            set_cell_value(schedule_sheet, f"{col}7", header)
        
        # フェーズとプロセスを書き込む
        row = 8
        for phase_idx, phase in enumerate(project.get_phases()):
            # フェーズID (アルファベット)
            phase_id = chr(65 + phase_idx + 1)  # A, B, C, ...
            
            # フェーズ行
            set_cell_value(schedule_sheet, f"B{row}", phase_id)
            set_cell_value(schedule_sheet, f"C{row}", phase.name)
            set_cell_value(schedule_sheet, f"E{row}", f"{phase.calculate_progress():.1f}%")
            
            if phase.get_start_date():
                set_cell_value(schedule_sheet, f"F{row}", phase.get_start_date())
            if phase.get_end_date():
                set_cell_value(schedule_sheet, f"G{row}", phase.get_end_date())
            
            row += 1
            
            # プロセス行
            for process_idx, process in enumerate(phase.get_processes()):
                process_id = f"{phase_id}{process_idx + 1}"
                
                set_cell_value(schedule_sheet, f"B{row}", process_id)
                set_cell_value(schedule_sheet, f"C{row}", process.name)
                set_cell_value(schedule_sheet, f"D{row}", process.assignee)
                set_cell_value(schedule_sheet, f"E{row}", f"{process.progress:.1f}%")
                
                if process.start_date:
                    set_cell_value(schedule_sheet, f"F{row}", process.start_date)
                if process.end_date:
                    set_cell_value(schedule_sheet, f"G{row}", process.end_date)
                
                set_cell_value(schedule_sheet, f"H{row}", process.estimated_hours)
                set_cell_value(schedule_sheet, f"I{row}", process.actual_hours)
                
                row += 1
    
    def _setup_input_sheet(self, workbook, project: Project):
        """
        入力シートを設定
        
        Args:
            workbook: Excelワークブック
            project: プロジェクト
        """
        input_sheet = workbook["入力"]
        
        # フェーズ名を設定 (I6:M6)
        phases = project.get_phases()
        for i, phase in enumerate(phases[:5]):  # 最大5フェーズまで
            col = chr(73 + i)  # I, J, K, ...
            set_cell_value(input_sheet, f"{col}6", phase.name)
        
        # プロセスを設定 (I7:M36)
        for i, phase in enumerate(phases[:5]):
            col = chr(73 + i)  # I, J, K, ...
            processes = phase.get_processes()
            
            for j, process in enumerate(processes[:30]):  # 最大30プロセスまで
                row = 7 + j
                set_cell_value(input_sheet, f"{col}{row}", process.name)
    
    def _create_task_sheets(self, workbook, project: Project):
        """
        各プロセスのタスクシートを作成
        
        Args:
            workbook: Excelワークブック
            project: プロジェクト
        """
        # 既存のプロセスシートを削除（最初の2シートを除く）
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names[2:]:
            del workbook[sheet_name]
        
        # プロセスごとにシートを作成
        for phase in project.get_phases():
            for process_idx, process in enumerate(phase.get_processes()):
                # シート名をプロセスIDとして設定
                phase_id = chr(65 + list(project.get_phases()).index(phase) + 1)
                process_id = f"{phase_id}{process_idx + 1}"
                
                # シートが既に存在する場合は削除
                if process_id in workbook.sheetnames:
                    del workbook[process_id]
                
                # 新しいシートを作成
                task_sheet = workbook.create_sheet(process_id)
                
                # ヘッダー設定
                set_cell_value(task_sheet, "B3", f"プロセス: {process.name}")
                set_cell_value(task_sheet, "B4", f"担当者: {process.assignee}")
                
                # タスク一覧
                set_cell_value(task_sheet, "B5", "タスク一覧:")
                
                # タスクを書き込む
                for task_idx, task in enumerate(process.get_tasks()):
                    row = 6 + task_idx
                    if task_idx < 12:  # 最大12タスクまで
                        set_cell_value(task_sheet, f"B{row}", task.name)
                        set_cell_value(task_sheet, f"C{row}", task.status.value)
