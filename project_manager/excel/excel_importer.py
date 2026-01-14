# project_manager/excel/excel_importer.py の変更部分

"""
Excelインポーター
Excelファイルからプロジェクトデータを読み込む機能
"""
import os
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Project, Phase, Process, Task, ProjectStatus, TaskStatus
from .excel_utils import get_cell_value, find_last_row_with_data, parse_excel_date


class ExcelImporter:
    """
    Excelファイルからプロジェクトデータを読み込むクラス
    """
    
    def __init__(self):
        """インポーターの初期化"""
        self.format_type = 'auto'
    
    def set_format_type(self, format_type: str):
        """
        インポートするフォーマットタイプを設定
        
        Args:
            format_type: フォーマットタイプ ('auto', 'standard', 'ms_project', 'simple')
        """
        self.format_type = format_type
    
    def detect_format(self, workbook) -> str:
        """
        Excelファイルのフォーマットを検出
        
        Args:
            workbook: openpyxlのワークブック
            
        Returns:
            検出されたフォーマット名
        """
        sheet_names = workbook.sheetnames
        
        # 標準フォーマット
        if "スケジュール" in sheet_names and "入力" in sheet_names:
            return "standard"
        
        # MS Project類似フォーマット
        if any(name in sheet_names for name in ["Tasks", "タスク", "Task List"]):
            if any(name in sheet_names for name in ["Resources", "リソース"]):
                return "ms_project"
        
        # シンプルフォーマット
        if any(name in sheet_names for name in ["Project", "プロジェクト", "Tasks", "タスク"]):
            return "simple"
        
        # 標準フォーマットに類似
        if any("スケジュール" in s for s in sheet_names) or any("計画" in s for s in sheet_names):
            return "standard_like"
        
        # 不明な場合はシンプルフォーマットとして処理
        return "simple"
    
    def import_from_file(self, file_path: str) -> Optional[Project]:
        """
        Excelファイルからプロジェクトデータをインポート
        
        Args:
            file_path: Excelファイルのパス
            
        Returns:
            インポートされたプロジェクト、または None（失敗時）
        """
        try:
            print(f"ファイル読み込み開始: {file_path}")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # フォーマットタイプの処理
            format_type = self.format_type
            
            if format_type == 'auto':
                # 自動検出
                format_type = self.detect_format(workbook)
                print(f"検出されたフォーマット: {format_type}")
            
            # フォーマットに応じたインポート処理を実行
            if format_type == "standard":
                return self._import_standard_format(workbook)
            elif format_type == "ms_project":
                return self._import_ms_project_format(workbook)
            elif format_type == "simple":
                return self._import_simple_format(workbook)
            elif format_type == "standard_like":
                return self._import_standard_like_format(workbook)
            else:
                print(f"サポートされていないフォーマットです: {format_type}")
                return None
                
        except Exception as e:
            # 詳細なエラー情報をコンソールに出力
            import traceback
            print(f"Excelのインポート中に例外が発生: {str(e)}")
            print(traceback.format_exc())
            return None
    
    def _import_standard_format(self, workbook) -> Optional[Project]:
        """
        標準フォーマットのインポート処理
        
        Args:
            workbook: Excelワークブック
            
        Returns:
            インポートされたプロジェクト、または None（失敗時）
        """
        # 必要なシートが存在するか確認
        required_sheets = ["スケジュール", "入力"]
        for sheet_name in required_sheets:
            if sheet_name not in workbook.sheetnames:
                print(f"必要なシート '{sheet_name}' が見つかりません")
                return None
        
        # スケジュールシートからプロジェクト情報を取得
        schedule_sheet = workbook["スケジュール"]
        project_name = get_cell_value(schedule_sheet, "B1")
        
        print(f"プロジェクト名: {project_name}")
        if not project_name:
            # B1セルにない場合は他のセルを探す
            project_name = self._find_project_name(schedule_sheet)
            if not project_name:
                project_name = "Excelからインポートしたプロジェクト"
                print(f"プロジェクト名が見つからないため、デフォルト名を使用: {project_name}")
        
        # プロジェクトを作成
        project = Project(name=project_name, description="Excelからインポートされました")
        
        # フェーズとプロセスをスケジュールシートから読み込む
        print("フェーズとプロセスの読み込み開始")
        self._import_phases_and_processes(workbook, project)
        
        # タスクを各シートから読み込む
        print("タスクの読み込み開始")
        self._import_tasks(workbook, project)
        
        # プロジェクトの検証
        if not project.get_phases():
            print("フェーズが1つもインポートされませんでした")
            return None
            
        print(f"インポート完了: {project.name}, フェーズ数: {len(project.get_phases())}")
        return project
    
    def _import_ms_project_format(self, workbook) -> Optional[Project]:
        """
        MS Project類似フォーマットのインポート処理
        
        Args:
            workbook: Excelワークブック
            
        Returns:
            インポートされたプロジェクト、または None（失敗時）
        """
        # タスクシートを特定
        task_sheet_name = None
        for name in ["Tasks", "タスク", "Task List"]:
            if name in workbook.sheetnames:
                task_sheet_name = name
                break
        
        if not task_sheet_name:
            print("タスクシートが見つかりません")
            return None
        
        tasks_sheet = workbook[task_sheet_name]
        
        # プロジェクト名を探索
        project_name = self._find_project_name(tasks_sheet)
        if not project_name:
            project_name = "MS Projectからインポートしたプロジェクト"
            print(f"プロジェクト名が見つからないため、デフォルト名を使用: {project_name}")
        
        # プロジェクトを作成
        project = Project(name=project_name, description="MS Project形式からインポートされました")
        
        # ヘッダー行を探索
        header_row = self._find_header_row(tasks_sheet)
        if not header_row:
            print(f"{task_sheet_name}シートのヘッダー行が見つかりません")
            
            # デフォルトフェーズを作成してタスクをインポート
            phase = Phase(name="メインフェーズ")
            project.add_phase(phase)
            self._import_simple_tasks(project, phase, tasks_sheet)
            
            return project
        
        # カラムマッピングを取得
        column_indices = self._get_column_indices(tasks_sheet, header_row)
        
        # WBSレベルまたは階層情報の列を特定
        level_column = None
        for col in range(1, min(10, tasks_sheet.max_column + 1)):
            header = tasks_sheet.cell(row=header_row, column=col).value
            if header and any(keyword in str(header).lower() for keyword in ["level", "wbs", "階層", "インデント"]):
                level_column = col
                break
        
        # 階層を構築
        self._build_hierarchy_from_ms_project(project, tasks_sheet, header_row, column_indices, level_column)
        
        # プロジェクトの検証
        if not project.get_phases():
            print("フェーズが1つもインポートされませんでした")
            # デフォルトフェーズを作成
            phase = Phase(name="メインフェーズ")
            project.add_phase(phase)
            
            # すべてのタスクをこのフェーズに紐づける
            for row in range(header_row + 1, tasks_sheet.max_row + 1):
                name_value = None
                if 'name' in column_indices:
                    name_value = tasks_sheet.cell(row=row, column=column_indices['name']).value
                
                if not name_value:
                    # 名前列が特定できない場合は2列目を試す
                    name_value = tasks_sheet.cell(row=row, column=2).value
                
                if name_value:
                    # プロセスを作成
                    assignee = ""
                    if 'assignee' in column_indices:
                        assignee = tasks_sheet.cell(row=row, column=column_indices['assignee']).value or ""
                    
                    process = Process(name=str(name_value), assignee=str(assignee) if assignee else "")
                    
                    # 日付情報を設定
                    if 'start_date' in column_indices:
                        start_date = parse_excel_date(tasks_sheet.cell(row=row, column=column_indices['start_date']).value)
                        if start_date:
                            process.set_start_date(start_date)
                    
                    if 'end_date' in column_indices:
                        end_date = parse_excel_date(tasks_sheet.cell(row=row, column=column_indices['end_date']).value)
                        if end_date:
                            process.set_end_date(end_date)
                    
                    # 工数情報を設定
                    estimated_hours = 0.0
                    if 'estimated_hours' in column_indices:
                        value = tasks_sheet.cell(row=row, column=column_indices['estimated_hours']).value
                        if value and isinstance(value, (int, float)):
                            estimated_hours = float(value)
                    
                    actual_hours = 0.0
                    if 'actual_hours' in column_indices:
                        value = tasks_sheet.cell(row=row, column=column_indices['actual_hours']).value
                        if value and isinstance(value, (int, float)):
                            actual_hours = float(value)
                    
                    process.update_hours(estimated_hours, actual_hours)
                    
                    # フェーズに追加
                    phase.add_process(process)
        
        return project
    
    def _import_simple_format(self, workbook) -> Optional[Project]:
        """
        シンプルフォーマットのインポート処理
        
        Args:
            workbook: Excelワークブック
            
        Returns:
            インポートされたプロジェクト、または None（失敗時）
        """
        # プロジェクトシートまたはタスクシートを特定
        main_sheet = None
        for name in ["Project", "プロジェクト", "Tasks", "タスク", "Sheet1", "シート1"]:
            if name in workbook.sheetnames:
                main_sheet = workbook[name]
                break
        
        if not main_sheet:
            main_sheet = workbook.active  # アクティブシートを使用
        
        # プロジェクト名を探索
        project_name = self._find_project_name(main_sheet)
        if not project_name:
            project_name = "シンプルフォーマットからインポートしたプロジェクト"
            print(f"プロジェクト名が見つからないため、デフォルト名を使用: {project_name}")
        
        # プロジェクトを作成
        project = Project(name=project_name, description="シンプルフォーマットからインポートされました")
        
        # デフォルトフェーズを作成
        phase = Phase(name="メインフェーズ")
        project.add_phase(phase)
        
        # ヘッダー行を探索
        header_row = self._find_header_row(main_sheet)
        if header_row:
            # 構造化データのインポート
            column_indices = self._get_column_indices(main_sheet, header_row)
            self._import_structured_simple_tasks(project, phase, main_sheet, header_row, column_indices)
        else:
            # 非構造化データからの単純インポート
            self._import_simple_tasks(project, phase, main_sheet)
        
        return project
    
    def _import_standard_like_format(self, workbook) -> Optional[Project]:
        """
        標準フォーマットに類似したフォーマットのインポート処理
        
        Args:
            workbook: Excelワークブック
            
        Returns:
            インポートされたプロジェクト、または None（失敗時）
        """
        # スケジュールシート的なものを探す
        schedule_sheet = None
        for sheet_name in workbook.sheetnames:
            if any(keyword in sheet_name for keyword in ["スケジュール", "計画", "Schedule", "Plan"]):
                schedule_sheet = workbook[sheet_name]
                break
        
        if not schedule_sheet:
            # 見つからない場合は最初のシートを使用
            schedule_sheet = workbook.active
            print(f"スケジュールシートが見つからないため、シート '{schedule_sheet.title}' を使用します")
        
        # プロジェクト名を探索
        project_name = self._find_project_name(schedule_sheet)
        if not project_name:
            project_name = "スケジュールからインポートしたプロジェクト"
            print(f"プロジェクト名が見つからないため、デフォルト名を使用: {project_name}")
        
        # プロジェクトを作成
        project = Project(name=project_name, description="類似フォーマットからインポートされました")
        
        # ヘッダー行を探索
        header_row = self._find_header_row(schedule_sheet)
        if not header_row:
            print(f"シート '{schedule_sheet.title}' のヘッダー行が見つかりません")
            
            # デフォルトフェーズを作成してタスクをインポート
            phase = Phase(name="メインフェーズ")
            project.add_phase(phase)
            self._import_simple_tasks(project, phase, schedule_sheet)
            
            return project
        
        # カラムマッピングを取得
        column_indices = self._get_column_indices(schedule_sheet, header_row)
        
        # インデント列または階層を示す列を探す
        indent_column = None
        for col in range(1, min(10, schedule_sheet.max_column + 1)):
            header = schedule_sheet.cell(row=header_row, column=col).value
            if header and isinstance(header, str) and any(keyword in header.lower() for keyword in ["level", "wbs", "階層", "インデント", "id"]):
                indent_column = col
                break
        
        if indent_column:
            # 階層情報があればそれを使う
            self._build_hierarchy_from_standard_like(project, schedule_sheet, header_row, column_indices, indent_column)
        else:
            # 階層情報がない場合はフラットな構造で作成
            phase = Phase(name="メインフェーズ")
            project.add_phase(phase)
            self._import_structured_simple_tasks(project, phase, schedule_sheet, header_row, column_indices)
        
        return project
    
    def _find_project_name(self, sheet) -> Optional[str]:
        """
        シートからプロジェクト名を探索
        
        Args:
            sheet: ワークシート
            
        Returns:
            プロジェクト名、または None
        """
        # プロジェクト名を探す一般的なキーワード
        project_keywords = ["プロジェクト", "project", "title", "タイトル", "名前"]
        
        # 最初の10行、5列を探索
        for row in range(1, min(10, sheet.max_row + 1)):
            for col in range(1, min(5, sheet.max_column + 1)):
                label_cell = sheet.cell(row=row, column=col).value
                if label_cell and isinstance(label_cell, str):
                    label_lower = label_cell.lower()
                    
                    # キーワードを含むラベルを見つけた場合
                    if any(keyword in label_lower for keyword in project_keywords):
                        # 右隣のセルの値を確認
                        value_cell = sheet.cell(row=row, column=col + 1).value
                        if value_cell:
                            return str(value_cell)
        
        # ラベルが見つからない場合は、単純に最初の行の目立つセルを探す
        for row in range(1, min(5, sheet.max_row + 1)):
            for col in range(1, min(5, sheet.max_column + 1)):
                value = sheet.cell(row=row, column=col).value
                if value and isinstance(value, str) and len(value) > 3:
                    return value
        
        return None
    
    def _find_header_row(self, sheet) -> Optional[int]:
        """
        ヘッダー行を探索
        
        Args:
            sheet: ワークシート
            
        Returns:
            ヘッダー行の行番号、または None
        """
        header_keywords = ["名前", "タスク", "開始", "終了", "担当", "進捗", "task", "name", "start", "end", "resource", "progress"]
        
        # 最初の20行を探索
        for row in range(1, min(20, sheet.max_row + 1)):
            keywords_found = 0
            for col in range(1, min(15, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower()
                    for keyword in header_keywords:
                        if keyword in cell_lower:
                            keywords_found += 1
                            break
            
            # 3つ以上のキーワードが見つかった行をヘッダーとみなす
            if keywords_found >= 3:
                return row
        
        return None
    
    def _get_column_indices(self, sheet, header_row: int) -> Dict[str, int]:
        """
        ヘッダー行からカラムのインデックスを取得
        
        Args:
            sheet: ワークシート
            header_row: ヘッダー行の行番号
            
        Returns:
            カラム名と列インデックスのマッピング
        """
        column_indices = {}
        column_mappings = {
            "id": ["id", "番号", "no", "no.", "wbs", "code"],
            "name": ["name", "名前", "タスク", "task", "作業", "項目", "内容", "title"],
            "assignee": ["assignee", "担当", "担当者", "resource", "リソース", "責任者", "担当部署"],
            "progress": ["progress", "進捗", "進捗率", "%", "完了", "completion"],
            "start_date": ["start", "開始", "開始日", "着手", "着手日"],
            "end_date": ["end", "終了", "終了日", "期日", "期限", "納期", "完了日", "deadline", "due"],
            "estimated_hours": ["estimate", "予想", "見積", "工数", "計画時間", "予定時間"],
            "actual_hours": ["actual", "実績", "実工数", "実時間", "実作業時間"]
        }
        
        for col in range(1, min(20, sheet.max_column + 1)):
            cell_value = sheet.cell(row=header_row, column=col).value
            if not cell_value:
                continue
            
            if not isinstance(cell_value, str):
                cell_value = str(cell_value)
            
            cell_lower = cell_value.lower()
            
            for key, aliases in column_mappings.items():
                for alias in aliases:
                    if alias in cell_lower:
                        column_indices[key] = col
                        break
        
        return column_indices
    
    def _build_hierarchy_from_ms_project(self, project, sheet, header_row: int, column_indices: Dict[str, int], level_column: Optional[int]):
        """
        MS Project形式のシートから階層構造を構築
        
        Args:
            project: プロジェクト
            sheet: ワークシート
            header_row: ヘッダー行の行番号
            column_indices: カラムのインデックスマッピング
            level_column: 階層レベルの列番号
        """
        # フェーズとプロセスを保持する辞書
        phases = {}
        
        # レベルが常に整数として解釈できるとは限らないため、文字インデントも考慮
        def get_level(row):
            if level_column:
                level_value = sheet.cell(row=row, column=level_column).value
                if level_value is not None:
                    if isinstance(level_value, (int, float)):
                        return int(level_value)
                    elif isinstance(level_value, str):
                        # インデント数をレベルとして解釈
                        return level_value.count('\t') + level_value.count('    ')
            
            # レベル列がない場合はインデントでレベルを推測
            name_col = column_indices.get('name', 2)
            name_value = sheet.cell(row=row, column=name_col).value
            if name_value and isinstance(name_value, str):
                return name_value.count('\t') + name_value.count('    ')
            
            return 0
        
        # データ行を処理
        for row in range(header_row + 1, sheet.max_row + 1):
            name_value = None
            if 'name' in column_indices:
                name_value = sheet.cell(row=row, column=column_indices['name']).value
            
            if not name_value:
                continue
            
            name_str = str(name_value).strip()
            if not name_str:
                continue
            
            # レベルを取得
            level = get_level(row)
            
            if level == 0 or level == 1:
                # フェーズとして扱う
                phase = Phase(name=name_str)
                project.add_phase(phase)
                phases[level] = phase
            else:
                # プロセスとして扱う
                assignee = ""
                if 'assignee' in column_indices:
                    assignee_value = sheet.cell(row=row, column=column_indices['assignee']).value
                    assignee = str(assignee_value) if assignee_value else ""
                
                process = Process(name=name_str, assignee=assignee)
                
                # 親フェーズを特定
                parent_level = level - 1
                while parent_level > 0 and parent_level not in phases:
                    parent_level -= 1
                
                parent_phase = phases.get(parent_level, None)
                if not parent_phase:
                    # 親が見つからない場合はデフォルトフェーズを作成
                    parent_phase = Phase(name="未分類フェーズ")
                    project.add_phase(parent_phase)
                    phases[0] = parent_phase
                
                # 日付情報を設定
                if 'start_date' in column_indices:
                    start_date = parse_excel_date(sheet.cell(row=row, column=column_indices['start_date']).value)
                    if start_date:
                        process.set_start_date(start_date)
                
                if 'end_date' in column_indices:
                    end_date = parse_excel_date(sheet.cell(row=row, column=column_indices['end_date']).value)
                    if end_date:
                        process.set_end_date(end_date)
                
                # 工数情報を設定
                estimated_hours = 0.0
                if 'estimated_hours' in column_indices:
                    value = sheet.cell(row=row, column=column_indices['estimated_hours']).value
                    if value and isinstance(value, (int, float)):
                        estimated_hours = float(value)
                
                actual_hours = 0.0
                if 'actual_hours' in column_indices:
                    value = sheet.cell(row=row, column=column_indices['actual_hours']).value
                    if value and isinstance(value, (int, float)):
                        actual_hours = float(value)
                
                process.update_hours(estimated_hours, actual_hours)
                
                # 進捗情報を設定
                if 'progress' in column_indices:
                    progress_value = sheet.cell(row=row, column=column_indices['progress']).value
                    if progress_value and isinstance(progress_value, (int, float)):
                        # プロセスに進捗率を直接設定するため、ダミーのタスクを作成
                        task = Task(name=f"{name_str} (自動生成)", status=TaskStatus.COMPLETED)
                        process.add_task(task)
                        
                        # 進捗に応じて未完了タスクも追加
                        if progress_value < 100:
                            remaining_task = Task(name=f"{name_str} (残り作業)", status=TaskStatus.NOT_STARTED)
                            process.add_task(remaining_task)
                
                # フェーズに追加
                parent_phase.add_process(process)
    
    def _build_hierarchy_from_standard_like(self, project, sheet, header_row: int, column_indices: Dict[str, int], indent_column: int):
        """
        標準フォーマット類似のシートから階層構造を構築
        
        Args:
            project: プロジェクト
            sheet: ワークシート
            header_row: ヘッダー行の行番号
            column_indices: カラムのインデックスマッピング
            indent_column: インデントの列番号
        """
        # 現在のフェーズを保持
        current_phase = None
        
        # インデント値またはIDの最後の文字に基づいて階層を判断
        for row in range(header_row + 1, sheet.max_row + 1):
            # ID列の値を取得
            id_value = sheet.cell(row=row, column=indent_column).value
            
            # 名前列の値を取得
            name_value = None
            if 'name' in column_indices:
                name_value = sheet.cell(row=row, column=column_indices['name']).value
            else:
                # 名前列が特定できない場合は2列目を試す
                name_value = sheet.cell(row=row, column=2).value
            
            if not id_value or not name_value:
                continue
            
            id_str = str(id_value)
            name_str = str(name_value)
            
            # IDがアルファベットのみならフェーズ、アルファベット+数字ならプロセス
            if id_str.isalpha():
                # フェーズ
                phase = Phase(name=name_str)
                project.add_phase(phase)
                current_phase = phase
                print(f"フェーズを追加: {name_str}, ID: {id_str}")
            elif any(c.isdigit() for c in id_str) and any(c.isalpha() for c in id_str):
                # プロセス
                if not current_phase:
                    # フェーズがない場合は作成
                    phase_name = "未分類フェーズ"
                    current_phase = Phase(name=phase_name)
                    project.add_phase(current_phase)
                    print(f"デフォルトフェーズを追加: {phase_name}")
                
                # プロセス情報を取得
                assignee = ""
                if 'assignee' in column_indices:
                    assignee_value = sheet.cell(row=row, column=column_indices['assignee']).value
                    assignee = str(assignee_value) if assignee_value else ""
                
                process = Process(name=name_str, assignee=assignee)
                
                # 日付情報を設定
                if 'start_date' in column_indices:
                    start_date = parse_excel_date(sheet.cell(row=row, column=column_indices['start_date']).value)
                    if start_date:
                        process.set_start_date(start_date)
                
                if 'end_date' in column_indices:
                    end_date = parse_excel_date(sheet.cell(row=row, column=column_indices['end_date']).value)
                    if end_date:
                        process.set_end_date(end_date)
                
                # 工数情報を設定
                estimated_hours = 0.0
                if 'estimated_hours' in column_indices:
                    value = sheet.cell(row=row, column=column_indices['estimated_hours']).value
                    try:
                        if value and isinstance(value, (int, float)):
                            estimated_hours = float(value)
                        elif value and isinstance(value, str) and value.replace('.', '', 1).isdigit():
                            estimated_hours = float(value)
                    except (ValueError, TypeError):
                        pass
                
                actual_hours = 0.0
                if 'actual_hours' in column_indices:
                    value = sheet.cell(row=row, column=column_indices['actual_hours']).value
                    try:
                        if value and isinstance(value, (int, float)):
                            actual_hours = float(value)
                        elif value and isinstance(value, str) and value.replace('.', '', 1).isdigit():
                            actual_hours = float(value)
                    except (ValueError, TypeError):
                        pass
                
                process.update_hours(estimated_hours, actual_hours)
                
                # フェーズに追加
                current_phase.add_process(process)
                print(f"プロセスを追加: {name_str}, ID: {id_str}, フェーズ: {current_phase.name}")
            else:
                # ID形式が想定外の場合
                print(f"警告: 不明なID形式です: {id_str}, 処理をスキップします。")
    
    def _import_structured_simple_tasks(self, project, phase, sheet, header_row: int, column_indices: Dict[str, int]):
        """
        構造化されたシンプルなタスク一覧をインポート
        
        Args:
            project: プロジェクト
            phase: フェーズ
            sheet: ワークシート
            header_row: ヘッダー行の行番号
            column_indices: カラムのインデックスマッピング
        """
        # データ行を処理
        for row in range(header_row + 1, sheet.max_row + 1):
            # 名前を取得
            name_value = None
            if 'name' in column_indices:
                name_value = sheet.cell(row=row, column=column_indices['name']).value
            
            if not name_value:
                continue
            
            name_str = str(name_value).strip()
            if not name_str:
                continue
            
            # プロセスを作成
            assignee = ""
            if 'assignee' in column_indices:
                assignee_value = sheet.cell(row=row, column=column_indices['assignee']).value
                if assignee_value:
                    assignee = str(assignee_value)
            
            process = Process(name=name_str, assignee=assignee)
            
            # 日付情報を設定
            if 'start_date' in column_indices:
                start_date = parse_excel_date(sheet.cell(row=row, column=column_indices['start_date']).value)
                if start_date:
                    process.set_start_date(start_date)
            
            if 'end_date' in column_indices:
                end_date = parse_excel_date(sheet.cell(row=row, column=column_indices['end_date']).value)
                if end_date:
                    process.set_end_date(end_date)
            
            # 工数情報を設定
            estimated_hours = 0.0
            if 'estimated_hours' in column_indices:
                value = sheet.cell(row=row, column=column_indices['estimated_hours']).value
                try:
                    if value and isinstance(value, (int, float)):
                        estimated_hours = float(value)
                    elif value and isinstance(value, str) and value.replace('.', '', 1).isdigit():
                        estimated_hours = float(value)
                except (ValueError, TypeError):
                    pass
            
            actual_hours = 0.0
            if 'actual_hours' in column_indices:
                value = sheet.cell(row=row, column=column_indices['actual_hours']).value
                try:
                    if value and isinstance(value, (int, float)):
                        actual_hours = float(value)
                    elif value and isinstance(value, str) and value.replace('.', '', 1).isdigit():
                        actual_hours = float(value)
                except (ValueError, TypeError):
                    pass
            
            process.update_hours(estimated_hours, actual_hours)
            
            # 進捗情報を設定
            if 'progress' in column_indices:
                progress_value = sheet.cell(row=row, column=column_indices['progress']).value
                try:
                    if progress_value is not None:
                        if isinstance(progress_value, (int, float)):
                            progress = float(progress_value)
                        elif isinstance(progress_value, str):
                            progress = float(progress_value.replace('%', ''))
                        
                        # プロセスに進捗率を直接設定するため、ダミーのタスクを作成
                        if progress > 0:
                            task = Task(name=f"{name_str} (完了部分)", status=TaskStatus.COMPLETED)
                            process.add_task(task)
                        
                        if progress < 100:
                            remaining_task = Task(name=f"{name_str} (残り部分)", status=TaskStatus.NOT_STARTED)
                            process.add_task(remaining_task)
                except (ValueError, TypeError):
                    pass
            
            # フェーズに追加
            phase.add_process(process)
    
    def _import_simple_tasks(self, project, phase, sheet):
        """
        単純なシートからタスクをインポート
        
        Args:
            project: プロジェクト
            phase: フェーズ
            sheet: ワークシート
        """
        # 各行を処理
        for row in range(1, sheet.max_row + 1):
            # 1列目と2列目のデータを取得
            col1 = sheet.cell(row=row, column=1).value
            col2 = sheet.cell(row=row, column=2).value
            
            # 有効なデータがあるか確認
            if col2 and isinstance(col2, str) and len(col2.strip()) > 0:
                # col2をタスク名として使用
                task_name = col2.strip()
                
                # col1が有効ならプロセス名、そうでなければタスク名をプロセス名として使用
                if col1 and isinstance(col1, str) and len(col1.strip()) > 0:
                    process_name = col1.strip()
                else:
                    process_name = task_name
                
                # プロセスを作成
                process = Process(name=process_name)
                
                # タスクを作成して追加
                task = Task(name=task_name)
                process.add_task(task)
                
                # フェーズに追加
                phase.add_process(process)
            elif col1 and isinstance(col1, str) and len(col1.strip()) > 0:
                # col1のみ有効な場合はプロセスとして追加
                process_name = col1.strip()
                process = Process(name=process_name)
                
                # デフォルトタスクを追加
                task = Task(name=f"{process_name} (デフォルトタスク)")
                process.add_task(task)
                
                # フェーズに追加
                phase.add_process(process)
    
    # 既存のメソッドはそのまま残します
    def _import_phases_and_processes(self, workbook, project: Project) -> None:
        """
        スケジュールシートからフェーズとプロセスをインポート
        
        Args:
            workbook: Excelワークブック
            project: プロジェクトオブジェクト
        """
        schedule_sheet = workbook["スケジュール"]
        
        # 入力シートからフェーズとプロセスの構造を解析
        input_sheet = workbook["入力"]
        
        # フェーズ名を取得 (I6:M6)
        phase_names = {}
        for col in range(9, 14):  # I=9, M=13
            cell_value = input_sheet.cell(row=6, column=col).value
            if cell_value:
                phase_names[col] = str(cell_value)
        
        # プロセスとフェーズの関係を取得 (I7:M36)
        process_to_phase = {}
        for col in range(9, 14):  # I=9, M=13
            if col not in phase_names:
                continue
                
            for row in range(7, 37):
                process_name = input_sheet.cell(row=row, column=col).value
                if process_name:
                    process_to_phase[str(process_name)] = phase_names[col]
        
        # スケジュールシートからフェーズとプロセスをインポート
        current_phases = {}  # フェーズID -> Phaseオブジェクト
        
        # 最後のデータ行を取得
        last_row = find_last_row_with_data(schedule_sheet, "B")
        
        for row in range(8, last_row + 1):
            id_value = get_cell_value(schedule_sheet, f"B{row}")
            name_value = get_cell_value(schedule_sheet, f"C{row}")
            
            if not id_value or not name_value:
                continue
            
            id_str = str(id_value)
            
            # IDの検証を強化
            print(f"処理中のID: {id_str}, 名前: {name_value}")
            
            # IDがアルファベットのみならフェーズ、アルファベット+数字ならプロセス
            if id_str.isalpha():
                # フェーズ
                phase = Phase(name=name_value)
                project.add_phase(phase)
                current_phases[id_str] = phase
                print(f"フェーズを追加: {name_value}, ID: {id_str}")
            elif any(c.isdigit() for c in id_str) and any(c.isalpha() for c in id_str):
                # プロセス
                assignee = get_cell_value(schedule_sheet, f"D{row}") or ""
                start_date = parse_excel_date(get_cell_value(schedule_sheet, f"F{row}"))
                end_date = parse_excel_date(get_cell_value(schedule_sheet, f"G{row}"))
                
                estimated_hours = get_cell_value(schedule_sheet, f"H{row}") or 0.0
                actual_hours = get_cell_value(schedule_sheet, f"I{row}") or 0.0
                
                # 数値型への変換を確実に行う
                try:
                    if not isinstance(estimated_hours, (int, float)):
                        estimated_hours = float(estimated_hours) if estimated_hours else 0.0
                    
                    if not isinstance(actual_hours, (int, float)):
                        actual_hours = float(actual_hours) if actual_hours else 0.0
                except (ValueError, TypeError):
                    # 変換できない場合はデフォルト値を使用
                    if not isinstance(estimated_hours, (int, float)):
                        estimated_hours = 0.0
                    if not isinstance(actual_hours, (int, float)):
                        actual_hours = 0.0
                
                # プロセスを作成 - ここで process 変数を定義
                process = Process(
                    name=name_value,
                    description="",
                    assignee=assignee
                )
                
                # 日付と工数を設定
                if start_date:
                    process.set_start_date(start_date)
                if end_date:
                    process.set_end_date(end_date)
                    
                process.update_hours(estimated_hours, actual_hours)
                
                # プロセスを適切なフェーズに追加
                # スケジュールシートのIDからフェーズを特定
                phase_id = ''.join([c for c in id_str if c.isalpha()])
                
                if phase_id in current_phases:
                    current_phases[phase_id].add_process(process)
                    print(f"プロセスを追加: {name_value}, ID: {id_str}, フェーズ: {phase_id}")
                else:
                    print(f"警告: フェーズID {phase_id} が見つかりません。プロセス {name_value} は追加されません。")
            else:
                # ID形式が想定外の場合
                print(f"警告: 不明なID形式です: {id_str}, 処理をスキップします。")

    def _import_tasks(self, workbook, project: Project) -> None:
        """
        各シートからタスクをインポート
        
        Args:
            workbook: Excelワークブック
            project: プロジェクトオブジェクト
        """
        # ExcelのプロセスIDとプロセスオブジェクトのマッピングを作成
        excel_process_map = {}
        
        # スケジュールシートを取得
        if "スケジュール" in workbook.sheetnames:
            schedule_sheet = workbook["スケジュール"]
            
            # 最後のデータ行を取得
            last_row = find_last_row_with_data(schedule_sheet, "B")
            
            # プロセスIDとマッピングを作成
            for row in range(8, last_row + 1):
                id_value = get_cell_value(schedule_sheet, f"B{row}")
                if not id_value:
                    continue
                
                id_str = str(id_value)
                
                # IDがアルファベット+数字の場合はプロセス
                if any(c.isdigit() for c in id_str) and any(c.isalpha() for c in id_str):
                    # プロセスIDからフェーズIDを抽出
                    phase_id = ''.join([c for c in id_str if c.isalpha()])
                    
                    # 対応するフェーズとプロセスを探す
                    for phase in project.get_phases():
                        for process in phase.get_processes():
                            # プロセス名がスケジュールシートの名前と一致するか確認
                            process_name = get_cell_value(schedule_sheet, f"C{row}")
                            if process.name == process_name:
                                # マッピングに追加
                                excel_process_map[id_str] = process
                                break
        
        print(f"タスクのインポート開始: プロセスIDとマッピング {excel_process_map}")
        
        # 各ExcelプロセスIDに対応するシートからタスクを読み込む
        for excel_process_id, process in excel_process_map.items():
            print(f"プロセスID {excel_process_id} のタスクを検索中...")
            
            # シート名がExcelプロセスIDに一致するものを探す
            if excel_process_id in workbook.sheetnames:
                sheet = workbook[excel_process_id]
                print(f"シート '{excel_process_id}' を処理中...")
                
                # タスク名を取得 (B6:B17)
                tasks_added = 0
                for row in range(6, 18):
                    task_name = get_cell_value(sheet, f"B{row}")
                    if not task_name:
                        continue
                    
                    # タスクステータスを取得
                    task_status_str = get_cell_value(sheet, f"C{row}")
                    task_status = TaskStatus.NOT_STARTED  # デフォルト
                    
                    # ステータス文字列をTaskStatus列挙体に変換
                    if task_status_str:
                        for status in TaskStatus:
                            if status.value == task_status_str:
                                task_status = status
                                break
                    
                    # タスクを作成してプロセスに追加
                    task = Task(name=str(task_name), status=task_status)
                    process.add_task(task)
                    tasks_added += 1
                    print(f"タスクを追加: {task_name}, 状態: {task_status.value}, プロセス: {process.name}")
                
                print(f"プロセス '{process.name}' に {tasks_added} 個のタスクを追加しました")
            else:
                print(f"警告: プロセスID '{excel_process_id}' に対応するシートが見つかりません")
                
